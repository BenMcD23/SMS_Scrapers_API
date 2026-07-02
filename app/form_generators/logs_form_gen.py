"""Fill the RAFAC Logs Form 202 template from a logs-form batch.

Live Demands: write Qty per demanded line, drop the lines that weren't
demanded. Cadet Nominal Roll: list the cadets the demand is for.
"""

import io
import os
import re
from collections import Counter
from datetime import date, timedelta

import openpyxl

TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)), "word_templates", "Logs Form 202.xlsx"
)

# item_type (app catalogue) -> Live Demands "Description" text, {size} substituted
DESCRIPTION_TEMPLATES: dict[str, str] = {
    "Beret":               "Beret RAF Size {size}",
    "Wedgewood Male":      "Shirt Long Sleeve Wedgewood Blue Male Size {size}",
    "Wedgewood Female":    "Shirt Long Sleeve Wedgewood Blue Female {size}",
    "Working Blue Male":   "Shirt Long Sleeve Mid Blue Male Size {size}",
    "Working Blue Female": "Shirt Long Sleeve Mid Blue Female {size}",
    "Jumper":              "Jumper Utility Blue Grey V-Neck (Unisex Item) Size {size}",
    "Trousers":            "Trousers Man's RAF No 2 Dress {size}",
    "Slacks":              "Trousers Woman's RAF {size}",
    "Skirts":              "Skirt RAF No 2 Dress {size}",
    "Tie":                 "Necktie Black (Unisex Item) {size}",  # size = Short|Standard
    "Belt":                "Belt Waist RAF (Unisex Item) Size 64-114cm",
}


def build_description(item_type: str, size: str) -> str | None:
    """Column-D description for an entry, or None if the item has no demand line."""
    template = DESCRIPTION_TEMPLATES.get(item_type)
    if not template:
        return None
    return template.format(size=size)


def _norm(value) -> str:
    # The template has typos ("Female115/42", trailing spaces, curly quotes) —
    # compare lowercased with all whitespace stripped and quotes unified.
    return re.sub(r"\s+", "", str(value)).replace("’", "'").lower()


def generate_logs_form(
    entries: list[tuple[str, str]],
    nominal_roll: list[tuple[str, str, str]],
) -> bytes:
    """entries: (item_type, size) per demanded item.
    nominal_roll: (rank, name, issue_or_exchange) per distinct person."""
    counts = Counter(
        desc for item_type, size in entries
        if (desc := build_description(item_type, size)) is not None
    )
    wanted = {_norm(desc): (desc, qty) for desc, qty in counts.items()}

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    rdd = date.today() + timedelta(days=21)

    ws = wb["Live Demands"]
    unmatched = dict(wanted)
    described_rows = [
        r for r in range(2, ws.max_row + 1) if ws.cell(r, 4).value not in (None, "")
    ]
    for r in described_rows:
        match = unmatched.pop(_norm(ws.cell(r, 4).value), None)
        if match:
            ws.cell(r, 5).value = match[1]
            ws.cell(r, 7).value = rdd
    # remove the demand lines that weren't ordered, bottom-up
    for r in reversed(described_rows):
        if ws.cell(r, 5).value in (None, ""):
            ws.delete_rows(r)
    # anything that didn't match a template line (template typos etc.) still
    # gets demanded — appended with the description for the QM to fix up
    next_row = 2
    while ws.cell(next_row, 4).value not in (None, ""):
        next_row += 1
    for desc, qty in unmatched.values():
        ws.cell(next_row, 4).value = desc
        ws.cell(next_row, 5).value = qty
        ws.cell(next_row, 6).value = "EA"
        ws.cell(next_row, 7).value = rdd
        ws.cell(next_row, 7).number_format = "D/M/YYYY"
        ws.cell(next_row, 8).value = "Routine"
        next_row += 1

    nr = wb["Cadet Nominal Roll"]
    for idx, (rank, name, issue) in enumerate(nominal_roll):
        row = 3 + idx  # row 3 holds the "A Example" placeholder
        nr.cell(row, 1).value = idx + 1
        nr.cell(row, 2).value = rank
        nr.cell(row, 3).value = name
        nr.cell(row, 4).value = issue
    for row in range(3 + len(nominal_roll), nr.max_row + 1):
        for col in range(1, 5):
            nr.cell(row, col).value = None

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


if __name__ == "__main__":
    # Self-check: every item_type/size combo from the app catalogue
    # (lib/stores-items.ts in the frontend) must match a Live Demands row.
    CATALOGUE_SIZES = {
        "Beret": ["48", "49", "50", "51", "52", "53", "54", "55", "56", "57", "58", "59", "60", "61", "62"],
        "Wedgewood Male": ["85/30", "90/33", "95/35", "95/36", "100/36", "100/38", "105/38", "105/39",
                           "110/39", "110/40", "115/40", "115/42", "120/42", "120/43", "125/43", "130/45", "135/48"],
        "Wedgewood Female": ["85/31", "85/33", "90/33", "90/35", "95/35", "95/36", "100/36", "100/38",
                             "105/38", "105/39", "110/39", "110/40", "115/40", "115/42", "120/42", "120/43"],
        "Working Blue Male": ["95/36", "100/38", "105/39", "110/40", "115/42", "120/43", "125/43", "130/45", "135/48"],
        "Working Blue Female": ["85/33", "90/35", "95/36", "100/38", "105/39", "110/40", "115/42", "120/43"],
        "Jumper": ["74", "82", "88", "94", "100", "106", "112", "118", "124", "130", "136"],
        "Trousers": ["66/68/76", "69/72/84", "72/72/88", "72/76/92", "72/80/96", "72/84/100", "72/88/104",
                     "75/72/88", "75/76/92", "75/80/96", "75/84/100", "75/88/104", "75/92/108",
                     "80/72/88", "80/76/92", "80/80/96", "80/84/100", "80/88/104", "80/92/108",
                     "80/96/112", "80/100/116", "80/104/120", "85/76/92", "85/80/96", "85/84/100",
                     "85/88/104", "85/92/108", "85/96/112", "85/100/116", "85/104/120", "85/108/124"],
        "Slacks": ["70/60/84", "70/65/89", "70/70/94", "70/75/99", "70/80/104", "75/60/84", "75/65/89",
                   "75/70/94", "75/75/99", "75/80/104", "75/85/109", "75/90/114", "75/95/119",
                   "80/65/89", "80/70/94", "80/75/99", "80/80/104", "80/85/109", "80/90/114",
                   "80/95/119", "80/100/124", "80/105/129", "85/65/89", "85/70/94", "85/75/99",
                   "85/80/104", "85/85/109", "85/90/114", "85/95/119", "85/100/124", "85/105/129"],
        "Skirts": ["65/60/84", "65/65/89", "65/70/94", "65/75/99", "65/80/104", "65/85/109", "65/90/114",
                   "65/95/119", "65/100/124", "65/105/129", "70/60/84", "70/65/89", "70/70/94",
                   "70/75/99", "70/80/104", "70/85/109", "70/90/114", "70/95/119", "70/100/124",
                   "70/105/129", "75/60/84", "75/65/89", "75/70/94", "75/75/99", "75/80/104",
                   "75/85/109", "75/90/114", "75/95/119", "75/100/124", "75/105/129"],
        "Tie": ["Short", "Standard"],
        "Belt": [""],
    }
    ws = openpyxl.load_workbook(TEMPLATE_PATH)["Live Demands"]
    template_rows = {
        _norm(ws.cell(r, 4).value)
        for r in range(2, ws.max_row + 1)
        if ws.cell(r, 4).value not in (None, "")
    }
    missing = [
        (item_type, size)
        for item_type, sizes in CATALOGUE_SIZES.items()
        for size in sizes
        if _norm(build_description(item_type, size)) not in template_rows
    ]
    # Known template typo: "Trousers Woman's RAF 80/70/99" where the catalogue has 80/75/99
    assert missing == [("Slacks", "80/75/99")], f"Unexpected unmatched combos: {missing}"
    entries = [("Beret", "56"), ("Beret", "56"), ("Tie", "Standard"), ("Slacks", "80/75/99")]
    out = generate_logs_form(entries, [("Cdt", "A Cadet", "Initial Issue")])
    check = openpyxl.load_workbook(io.BytesIO(out))["Live Demands"]
    rows = {check.cell(r, 4).value: check.cell(r, 5).value for r in range(2, check.max_row + 1)
            if check.cell(r, 4).value not in (None, "")}
    assert rows == {
        "Beret RAF Size 56": 2,
        "Necktie Black (Unisex Item) Standard": 1,
        "Trousers Woman's RAF 80/75/99": 1,  # appended — no template row
    }, f"Unexpected output rows: {rows}"
    rdds = {check.cell(r, 7).value.date() if hasattr(check.cell(r, 7).value, "date") else check.cell(r, 7).value
            for r in range(2, check.max_row + 1) if check.cell(r, 4).value not in (None, "")}
    assert rdds == {date.today() + timedelta(days=21)}, f"Unexpected RDDs: {rdds}"
    print("logs_form_gen self-check OK")
